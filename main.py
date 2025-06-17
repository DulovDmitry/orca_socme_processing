from db_handle import DataBase
from homo_lumo_gap_from_orca import HomoLumoReader
from socme import SocmeReader
from my_collections import *
from my_config import *

import logging
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill

database = DataBase()

logging.basicConfig(level=logging.INFO , format="%(asctime)s - [%(levelname)s] - %(funcName)s - %(message)s")

def fill_database_with_homo_lumo_data():
	logging.info("Filling the database with HOMO and LUMO data")
	homo_lumo_reader = HomoLumoReader()

	for filename in os.listdir(OUTFILES_DIR):
		if not ".out" in filename:
			continue

		filename_split = filename.split(".")[0].split("_")
		long_name = filename_split[0]
		if not long_name in NAMES_DICT:
			continue

		short_name = NAMES_DICT[long_name]

		functional = [value for value in filename_split if value in FUNCTIONALS]
		if len(functional) == 1:
			functional = functional[0]
		else:
			continue

		homo, lumo = homo_lumo_reader.find_homo_lumo_energies(OUTFILES_DIR + filename)
		if homo is None or lumo is None:
			continue

		database.add_homo_lumo_data(long_name, short_name, functional, homo, lumo)
	logging.info("The database has been filled with HOMO and LUMO data")


def fill_excel_with_homo_lumo_data():
	logging.info("Writing HOMO and LUMO data to the excel file")
	for functional in FUNCTIONALS:
		for parameter in HOMO_LUMO_PARAMETERS:
			res = database.get_homo_lumo_data(functional, parameter)
			df = get_dataframe_from_list_of_parameters(res)

			if len(df) == 0:
				continue

			sheet_name = prepare_excel_sheet(functional, parameter)

			with pd.ExcelWriter(EXCEL_FILE_NAME, mode='a', if_sheet_exists='overlay') as writer:
				df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=4, startcol=2)
	logging.info("HOMO and LUMO data has been written to the excel file")


def fill_database_with_singlet_triplet_data():
	logging.info("Filling the database with singlet and triplet data")
	for filename in os.listdir(OUTFILES_DIR):
		if not ".out" in filename:
			continue

		if not ("SOC" in filename and "pbe0" in filename):
			continue

		filename_split = filename.split(".")[0].split("_")
		long_name = filename_split[0]
		if not long_name in NAMES_DICT:
			continue

		short_name = NAMES_DICT[long_name]

		socme_reader = SocmeReader(OUTFILES_DIR + filename)
		database.add_singlet_triplet_data(long_name, short_name, "pbe0", socme_reader.S1_energy, socme_reader.T1_energy, socme_reader.delta_E_S1_T1)
	logging.info("The database has been filled with singlet and triplet data")


def fill_excel_with_singlet_triplet_data():
	logging.info("Writing singlet and triplet data to the excel file")
	for functional in FUNCTIONALS:
		for parameter in SINGLET_TRIPLET_PARAMETERS:
			res = database.get_singlet_triplet_data(functional, parameter)
			df = get_dataframe_from_list_of_parameters(res)

			if len(df) == 0:
				continue

			sheet_name = prepare_excel_sheet(functional, parameter)

			with pd.ExcelWriter(EXCEL_FILE_NAME, mode='a', if_sheet_exists='overlay') as writer:
				df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=4, startcol=2)
	logging.info("singlet and triplet data has been written to the excel file")


def fill_database_with_socme_data():
	logging.info("Filling the database with SOCME data")
	for filename in os.listdir(OUTFILES_DIR):
		if not ".out" in filename:
			continue

		if not ("SOC" in filename and "pbe0" in filename):
			continue

		filename_split = filename.split(".")[0].split("_")
		long_name = filename_split[0]
		if not long_name in NAMES_DICT:
			continue

		short_name = NAMES_DICT[long_name]

		socme_reader = SocmeReader(OUTFILES_DIR + filename)
		database.add_socme_data(long_name, short_name, "pbe0", socme_reader.T1_S1_SOCME, socme_reader.T1_S2_SOCME)

		# построение результирующего графика и сохранение в формате pdf
		socme_reader.create_summary_plot(short_name)
		socme_reader.save_summary_plot_as_pdf(short_name)
	logging.info("The database has been filled with SOCME data")


def fill_excel_with_socme_data():
	logging.info("Writing SOCME data to the excel file")
	for functional in FUNCTIONALS:
		for parameter in SOCME_PARAMETERS:
			res = database.get_socme_data(functional, parameter)
			df = get_dataframe_from_list_of_parameters(res)

			if len(df) == 0:
				continue

			sheet_name = prepare_excel_sheet(functional, parameter)

			with pd.ExcelWriter(EXCEL_FILE_NAME, mode='a', if_sheet_exists='overlay') as writer:
				df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=4, startcol=2)
	logging.info("SOCME data has been written to the excel file")


def get_dataframe_from_list_of_parameters(list_of_parameters):
	df = pd.DataFrame(list_of_parameters, columns=['Key', 'Value'])
	df['Letter'] = df['Key'].str[0]
	df['Number'] = df['Key'].str[1]
	pivot_df = df.pivot(index='Number', columns='Letter', values='Value').sort_index().sort_index(axis=1)
	return pivot_df


def check_orca_version():
	logging.info("Staring ORCA version check")
	for filename in os.listdir(OUTFILES_DIR):
		if not ".out" in filename:
			continue

		with open(OUTFILES_DIR + filename, 'r') as file:
			lines = file.readlines()
			if not ('6.0.0' in lines[53] or '6.0.0' in lines[60]):
				logging.warning(f"ORCA version is not 6.0.0 in the file {filename}")
	logging.info("ORCA version check done!")


def check_geometry_convergence():
	logging.info("Staring geometry convergence check")
	for filename in os.listdir(OUTFILES_DIR):
		if not ".out" in filename:
			continue

		if not "opt" in filename:
			continue

		flag = False
		with open(OUTFILES_DIR + filename, 'r') as file:
			lines = file.readlines()

			for line in lines:
				if "OPTIMIZATION HAS CONVERGED" in line:
					flag = True
					break

			if not flag:
				logging.warning(f"Geometry did not converged: {filename}")
	logging.info("Geometry convergence check done!")


def prepare_excel_sheet(functional, parameter):
	sheet_name = f'{functional} {parameter}'
	wb = load_workbook(EXCEL_FILE_NAME)
	if not sheet_name in wb.sheetnames:
		new_sheet = wb.copy_worksheet(wb['template'])
		new_sheet.title = sheet_name
		title_cell = new_sheet['B2']
		title_cell.value = PARAMETERS_DICT[parameter] + " (" + FUNCTIONALS_DICT[functional] + ")"
		if parameter in CELL_COLORS_DICT.keys():
			title_cell.fill = PatternFill(fill_type='solid',
										  start_color=CELL_COLORS_DICT[parameter],
										  end_color=CELL_COLORS_DICT[parameter])
	wb.save(EXCEL_FILE_NAME)
	return sheet_name


def main():
	check_orca_version()
	check_geometry_convergence()
	fill_database_with_homo_lumo_data()
	fill_excel_with_homo_lumo_data()
	fill_database_with_singlet_triplet_data()
	fill_excel_with_singlet_triplet_data()
	fill_database_with_socme_data()
	fill_excel_with_socme_data()

if __name__ == "__main__":
	main()
